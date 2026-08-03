import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.db.models import Admin, MovieCode, MovieCodeAudit, MovieCodeStatus
from app.services.bulk_import import (
    ParsedRow,
    apply_import_plan,
    build_import_plan,
    build_template_xlsx,
    parse_csv,
    parse_xlsx,
)


async def _admin(session_factory) -> None:
    async with session_factory() as session:
        session.add(Admin(telegram_id=800))
        await session.commit()


def test_parse_csv_and_reports_invalid_rows():
    rows = parse_csv(b"operation, code, title\nupload, 00123, Movie\ndelete, 00001,\n")
    assert [(row.operation, row.code, row.title, row.row_number) for row in rows] == [
        ("upload", "00123", "Movie", 2),
        ("delete", "00001", None, 3),
    ]
    invalid = parse_csv(b"operation,code,title\nunknown,123,x\nupload,,x\n")
    assert "Строка 2" in invalid.parse_errors[0]
    assert "Строка 3" in invalid.parse_errors[1]


def test_parse_xlsx_marks_numeric_code_as_risky():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["operation", "code", "title"])
    sheet.append(["upload", 123, "Movie"])
    content = io.BytesIO()
    workbook.save(content)
    rows = parse_xlsx(content.getvalue())
    assert rows[0].code == "123"
    assert rows[0].code_looks_numeric_risk is True


def test_template_xlsx_contains_delete_example():
    workbook = load_workbook(io.BytesIO(build_template_xlsx()), data_only=True)
    try:
        assert any(row[0] == "delete" for row in workbook.active.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_plan_and_apply_bulk_import(session_factory):
    await _admin(session_factory)
    async with session_factory() as session:
        session.add_all([
            MovieCode(code="same", title="Same", status=MovieCodeStatus.ACTIVE),
            MovieCode(code="conflict", title="Old", status=MovieCodeStatus.ACTIVE),
            MovieCode(code="inactive", title="Old inactive", status=MovieCodeStatus.INACTIVE),
            MovieCode(code="delete", title="Delete me", status=MovieCodeStatus.ACTIVE),
        ])
        await session.commit()
        plan = await build_import_plan(session, [
            ParsedRow("upload", "new", "New", False, 2),
            ParsedRow("upload", "same", "Same", False, 3),
            ParsedRow("upload", "conflict", "Replacement", False, 4),
            ParsedRow("upload", "inactive", "Restored", False, 5),
            ParsedRow("delete", "delete", None, False, 6),
            ParsedRow("delete", "missing", None, False, 7),
        ])
    assert [row.code for row in plan.to_create] == ["new"]
    assert [row.code for row in plan.to_restore] == ["inactive"]
    assert [row.code for row in plan.unchanged] == ["same"]
    assert [row.code for row, _ in plan.conflicts] == ["conflict"]
    assert [row.code for row in plan.to_delete] == ["delete"]
    assert [row.code for row in plan.delete_skipped] == ["missing"]

    async with session_factory() as session:
        summary = await apply_import_plan(session, plan, 800, "keep")
    assert summary.created == 2
    assert summary.skipped == 3
    async with session_factory() as session:
        assert (await session.get(MovieCode, "conflict")).title == "Old"
        replacement_plan = await build_import_plan(session, [
            ParsedRow("upload", "conflict", "Replacement", False, 4),
            ParsedRow("upload", "inactive", "Restored", False, 5),
        ])
        summary = await apply_import_plan(session, replacement_plan, 800, "replace")
    assert summary.updated == 1
    async with session_factory() as session:
        inactive = await session.get(MovieCode, "inactive")
        deleted = await session.get(MovieCode, "delete")
        audits = list((await session.scalars(select(MovieCodeAudit))).all())
    assert inactive.status is MovieCodeStatus.ACTIVE
    assert inactive.title == "Restored"
    assert deleted.status is MovieCodeStatus.INACTIVE
    assert all(audit.source.value == "bulk_import" for audit in audits)


@pytest.mark.asyncio
async def test_bulk_upload_restores_inactive_codes(session_factory):
    await _admin(session_factory)
    async with session_factory() as session:
        session.add_all([
            MovieCode(code="same-title", title="Original", status=MovieCodeStatus.INACTIVE),
            MovieCode(code="new-title", title="Original", status=MovieCodeStatus.INACTIVE),
        ])
        await session.commit()
        plan = await build_import_plan(session, [
            ParsedRow("upload", "same-title", "Original", False, 2),
            ParsedRow("upload", "new-title", "Renamed", False, 3),
        ])

    assert [row.code for row in plan.to_restore] == ["same-title", "new-title"]
    assert not plan.unchanged
    assert not plan.conflicts

    async with session_factory() as session:
        summary = await apply_import_plan(session, plan, 800, "keep")
    assert summary.created == 2

    async with session_factory() as session:
        same_title = await session.get(MovieCode, "same-title")
        new_title = await session.get(MovieCode, "new-title")
        audits = list((await session.scalars(
            select(MovieCodeAudit).where(MovieCodeAudit.code.in_(["same-title", "new-title"]))
        )).all())
    assert same_title.status is MovieCodeStatus.ACTIVE
    assert same_title.title == "Original"
    assert new_title.status is MovieCodeStatus.ACTIVE
    assert new_title.title == "Renamed"
    assert {(audit.code, audit.action, audit.source.value) for audit in audits} == {
        ("same-title", "restore", "bulk_import"),
        ("new-title", "restore", "bulk_import"),
    }
