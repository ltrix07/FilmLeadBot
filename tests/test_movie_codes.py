import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.db.models import Admin, MovieCode, MovieCodeAudit, MovieCodeStatus
from app.services.movie_codes import (
    build_active_codes_export_xlsx,
    create_code,
    deactivate_code,
    restore_code,
    update_title,
)


async def _admin(session_factory, telegram_id: int = 800) -> None:
    async with session_factory() as session:
        session.add(Admin(telegram_id=telegram_id))
        await session.commit()


@pytest.mark.asyncio
async def test_create_code_records_audit_and_rejects_duplicates(session_factory):
    await _admin(session_factory)
    async with session_factory() as session:
        created = await create_code(session, "00123", "Original", 800)
        assert created.status is MovieCodeStatus.ACTIVE
    async with session_factory() as session:
        audit = await session.scalar(select(MovieCodeAudit).where(MovieCodeAudit.code == "00123"))
        assert audit.action.value == "create"
        assert audit.old_title is None
        assert audit.new_title == "Original"
        with pytest.raises(ValueError):
            await create_code(session, "00123", "Duplicate", 800)


@pytest.mark.asyncio
async def test_update_and_deactivate_are_audited_and_idempotent(session_factory):
    await _admin(session_factory)
    async with session_factory() as session:
        await create_code(session, "00123", "Original", 800)
        updated = await update_title(session, "00123", "Updated", 800)
        assert updated.code == "00123"
        assert updated.title == "Updated"
        await deactivate_code(session, "00123", 800)
    async with session_factory() as session:
        count_before = await session.scalar(
            select(func.count()).select_from(MovieCodeAudit).where(MovieCodeAudit.code == "00123")
        )
        inactive = await deactivate_code(session, "00123", 800)
        assert inactive.status is MovieCodeStatus.INACTIVE
        count_after = await session.scalar(
            select(func.count()).select_from(MovieCodeAudit).where(MovieCodeAudit.code == "00123")
        )
        audits = list((await session.scalars(
            select(MovieCodeAudit).where(MovieCodeAudit.code == "00123").order_by(MovieCodeAudit.id)
        )).all())
    assert count_after == count_before == 3
    assert audits[1].action.value == "update"
    assert (audits[1].old_title, audits[1].new_title) == ("Original", "Updated")
    assert audits[2].action.value == "deactivate"


@pytest.mark.asyncio
async def test_restore_preserves_or_replaces_title_and_rejects_active_code(session_factory):
    await _admin(session_factory)
    async with session_factory() as session:
        await create_code(session, "00123", "Original", 800)
        await deactivate_code(session, "00123", 800)
        restored = await restore_code(session, "00123", 800, None)
        assert restored.status is MovieCodeStatus.ACTIVE
        assert restored.title == "Original"
        with pytest.raises(ValueError):
            await restore_code(session, "00123", 800, "Ignored")
        await deactivate_code(session, "00123", 800)
        restored = await restore_code(session, "00123", 800, "Replacement")
        assert restored.title == "Replacement"
    async with session_factory() as session:
        code = await session.get(MovieCode, "00123")
        audits = list((await session.scalars(
            select(MovieCodeAudit).where(MovieCodeAudit.code == "00123").order_by(MovieCodeAudit.id)
        )).all())
    assert code.code == "00123"
    assert code.title == "Replacement"
    assert audits[-1].action.value == "restore"
    assert (audits[-1].old_title, audits[-1].new_title) == ("Original", "Replacement")


@pytest.mark.asyncio
async def test_active_codes_export_includes_only_active_codes_and_preserves_text(session_factory):
    await _admin(session_factory)
    async with session_factory() as session:
        await create_code(session, "00042", "Leading zero", 800)
        await create_code(session, "123", "Active", 800)
        await create_code(session, "inactive", "Hidden", 800)
        await deactivate_code(session, "inactive", 800)
        content = await build_active_codes_export_xlsx(session)

    import io
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows == [("code", "title"), ("00042", "Leading zero"), ("123", "Active")]
    assert worksheet["A2"].number_format == "@"
    workbook.close()
