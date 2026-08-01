from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import MovieCode, MovieCodeStatus
from app.services.movie_codes import create_code, deactivate_code, restore_code, update_title


@dataclass(frozen=True)
class ParsedRow:
    operation: str
    code: str
    title: str | None
    code_looks_numeric_risk: bool
    row_number: int


class ParsedRows(list[ParsedRow]):
    """Parsed rows together with validation errors, while remaining list-compatible."""

    def __init__(self, rows: list[ParsedRow] | None = None, parse_errors: list[str] | None = None) -> None:
        super().__init__(rows or [])
        self.parse_errors = parse_errors or []


@dataclass(frozen=True)
class ImportPlan:
    to_create: list[ParsedRow]
    to_restore: list[ParsedRow]
    unchanged: list[ParsedRow]
    conflicts: list[tuple[ParsedRow, str]]
    to_delete: list[ParsedRow]
    delete_skipped: list[ParsedRow]
    parse_errors: list[str]
    numeric_risk_codes: list[str]


@dataclass(frozen=True)
class ImportSummary:
    created: int
    updated: int
    deactivated: int
    skipped: int


_EXPECTED_HEADERS = {"operation", "code", "title"}


def _header_mapping(headers: list[object] | tuple[object, ...] | None) -> tuple[dict[str, int], list[str]]:
    normalized = [str(value or "").strip().lower() for value in (headers or [])]
    mapping = {value: index for index, value in enumerate(normalized)}
    if set(normalized) != _EXPECTED_HEADERS or len(normalized) != len(_EXPECTED_HEADERS):
        return {}, ["Некорректный заголовок: ожидаются колонки operation, code, title."]
    return mapping, []


def _parse_values(values: list[object], mapping: dict[str, int], row_number: int, *, numeric_risk: bool) -> tuple[ParsedRow | None, list[str]]:
    def value(name: str) -> object:
        index = mapping[name]
        return values[index] if index < len(values) else None

    operation = str(value("operation") or "").strip().lower()
    code_value = value("code")
    if numeric_risk:
        code = str(int(code_value))
    else:
        code = str(code_value or "").strip()
    title_value = value("title")
    title = str(title_value or "").strip()
    errors: list[str] = []
    if operation not in {"upload", "delete"}:
        errors.append(f"Строка {row_number}: неизвестная операция '{operation}'")
    if not code:
        errors.append(f"Строка {row_number}: пустой код")
    if operation == "upload" and not title:
        errors.append(f"Строка {row_number}: пустое название")
    if errors:
        return None, errors
    return ParsedRow(operation, code, title if operation == "upload" else None, numeric_risk, row_number), []


def parse_csv(content: bytes) -> ParsedRows:
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    mapping, errors = _header_mapping(reader.fieldnames)
    if errors:
        return ParsedRows(parse_errors=errors)
    original_headers = {header.strip().lower(): header for header in reader.fieldnames or []}
    rows: list[ParsedRow] = []
    for row_number, row_data in enumerate(reader, start=2):
        values = [row_data.get(original_headers[name]) for name in ("operation", "code", "title")]
        if not any(str(value or "").strip() for value in values):
            continue
        row, row_errors = _parse_values(
            values, {"operation": 0, "code": 1, "title": 2}, row_number, numeric_risk=False
        )
        errors.extend(row_errors)
        if row is not None:
            rows.append(row)
    return ParsedRows(rows, errors)


def parse_xlsx(content: bytes) -> ParsedRows:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        mapping, errors = _header_mapping(headers)
        if errors:
            return ParsedRows(parse_errors=errors)
        rows: list[ParsedRow] = []
        for row_number, values in enumerate(iterator, start=2):
            values_list = list(values)
            if not any(value is not None and str(value).strip() for value in values_list):
                continue
            code_value = values_list[mapping["code"]] if mapping["code"] < len(values_list) else None
            numeric_risk = isinstance(code_value, (int, float)) and not isinstance(code_value, bool)
            row, row_errors = _parse_values(values_list, mapping, row_number, numeric_risk=numeric_risk)
            errors.extend(row_errors)
            if row is not None:
                rows.append(row)
        return ParsedRows(rows, errors)
    finally:
        # ``load_workbook`` may fail before assignment; close only if it succeeded.
        if "workbook" in locals():
            workbook.close()


async def build_import_plan(session: AsyncSession, rows: list[ParsedRow]) -> ImportPlan:
    to_create: list[ParsedRow] = []
    to_restore: list[ParsedRow] = []
    unchanged: list[ParsedRow] = []
    conflicts: list[tuple[ParsedRow, str]] = []
    to_delete: list[ParsedRow] = []
    delete_skipped: list[ParsedRow] = []
    for row in rows:
        movie_code = await session.get(MovieCode, row.code)
        if row.operation == "upload":
            if movie_code is None:
                to_create.append(row)
            elif movie_code.status is not MovieCodeStatus.ACTIVE:
                to_restore.append(row)
            elif movie_code.title == row.title:
                unchanged.append(row)
            else:
                conflicts.append((row, movie_code.title))
        elif movie_code is None or movie_code.status is not MovieCodeStatus.ACTIVE:
            delete_skipped.append(row)
        else:
            to_delete.append(row)
    return ImportPlan(
        to_create, to_restore, unchanged, conflicts, to_delete, delete_skipped,
        list(getattr(rows, "parse_errors", [])),
        [row.code for row in rows if row.code_looks_numeric_risk],
    )


async def apply_import_plan(
    session: AsyncSession,
    plan: ImportPlan,
    admin_telegram_id: int,
    conflict_resolution: Literal["replace", "keep"],
) -> ImportSummary:
    for row in plan.to_create:
        await create_code(session, row.code, row.title or "", admin_telegram_id, source="bulk_import")
    for row in plan.to_restore:
        await restore_code(session, row.code, admin_telegram_id, row.title, source="bulk_import")
    updated = 0
    if conflict_resolution == "replace":
        for row, _ in plan.conflicts:
            current = await session.get(MovieCode, row.code)
            if current is not None and current.status is MovieCodeStatus.ACTIVE:
                await update_title(session, row.code, row.title or "", admin_telegram_id, source="bulk_import")
            else:
                await restore_code(session, row.code, admin_telegram_id, row.title, source="bulk_import")
            updated += 1
    for row in plan.to_delete:
        await deactivate_code(session, row.code, admin_telegram_id, source="bulk_import")
    return ImportSummary(
        created=len(plan.to_create) + len(plan.to_restore), updated=updated, deactivated=len(plan.to_delete),
        skipped=len(plan.unchanged) + len(plan.delete_skipped) + (len(plan.conflicts) if conflict_resolution == "keep" else 0),
    )


def build_template_xlsx() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["operation", "code", "title"])
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.append(["upload", "00123", "Пример названия"])
    for cell in worksheet["B"]:
        cell.number_format = "@"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def build_conflicts_report_csv(plan: ImportPlan) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["code", "current_title", "new_title"])
    for row, current_title in plan.conflicts:
        writer.writerow([row.code, current_title, row.title or ""])
    return output.getvalue().encode("utf-8-sig")
